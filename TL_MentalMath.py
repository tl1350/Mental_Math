#TL_MentalMath.py
#Started 2021.12.04
'''
Purpose: Drill to skill. Help develop mental math skills through the below drills: 

    1 Addtion (and Subtraction) - 3 Digits >>>>>>> strategy figure out left to right
    2 Multiplication - 2 digits >>>>>>>>>>>> strategy addition way, factoring
    3 Squares - 2 digits >>>>>>>> stategy, figure out a^2 , utilize (a+b)(a-b) = a^2 - b^2

    Answers are logged to Excel

'''

import random # for generating random numbers
import pyinputplus as pyip #for dealing with inputs https://pyinputplus.readthedocs.io/en/latest/
import time #for keeping track of time (might use another module later)
import logging # for debugging
from datetime import datetime  #https://docs.python.org/3/library/datetime.html
import openpyxl #https://openpyxl.readthedocs.io/en/stable/
import copy
import os #https://docs.python.org/3/library/os.html
import math_functions
import pyttsx3 #https://pyttsx3.readthedocs.io/en/latest/engine.html to speak out problems

excel_file = "C:/Users/tonya/Downloads/MentalMathResult.xlsx" # change your excel location here

def menu(wb, sheet):
    

    start_time = time.time() # keeps track of start type
    num_right = 0 #numbers of questions answered correctly

    #inputs
    program_type = pyip.inputMenu(['add', 'mult', 'squares'], numbered=True) #Choose a program
    num_problems = pyip.inputInt(prompt='Enter the number of practice problems: ',min=1,max=20) #choose number of problems
    

    for problem_id in range(1,num_problems+1):
        
        if program_type == "add":
            info = math_functions.add_subt(problem_id)  
        elif program_type == "mult":
            info = math_functions.mult(problem_id, program_type)
        elif program_type == "squares":
            info = math_functions.mult(problem_id, program_type)
        
        num_right += info["point"]
        # Log results to Excel
        sheet.cell(row=sheet.max_row + 1, column=1).value = problem_id
        sheet.cell(row=sheet.max_row, column=2).value = info["num1"]
        sheet.cell(row=sheet.max_row, column=3).value = info["num2"]
        sheet.cell(row=sheet.max_row, column=4).value = info["answer"]
        sheet.cell(row=sheet.max_row, column=5).value = info["type"]
        sheet.cell(row=sheet.max_row, column=6).value = info["point"]
        sheet.cell(row=sheet.max_row, column=7).value = info["formula"]
    
    wb.save(excel_file)
    #add in function here that shows results
    #time_stats = datetime.now() - start_time
    time_stats = time.time() - start_time
    
    print('Stats')
    print(f'Score: {num_right}/{problem_id} | {float(num_right)/problem_id:.0%}')
    print(f'Time: {round(time_stats,0)} seconds')


    #Would you like to pracice for?
    replay_option = pyip.inputYesNo(prompt="Do you want to play again? ")
   
    if replay_option == "yes":
        menu(wb, sheet)
    else:
        os.system("start EXCEL.EXE " + excel_file)
        replay_option = pyip.inputYesNo(prompt="Press Y to exit:  ")


#### Running the Program ###########
def main():

    print('Procedural Learning: Drill to Skill - Mental Math')

    #Open Excel to save the results
    try:
        wb = openpyxl.load_workbook(filename = 'C:/Users/tonya/Downloads/MentalMathResult.xlsx')
    except:
        wb = openpyxl.Workbook()

    #create sheet name using timestamp
    session = datetime.now().strftime("%m.%d--%H.%M.%S")
    sheet = wb.create_sheet(title = session,index = 0)
    wb.active = 0

    #creates headers
    sheet["A1"] = " ID "
    sheet["B1"] = "Num1"
    sheet["C1"] = "Num2"
    sheet["D1"] = "Answer"
    sheet["E1"] = "Type"
    sheet["F1"] = "Point"
    sheet["G1"] = "Formula"

    menu(wb,sheet)

main()
